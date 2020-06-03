#import sys
from tkinter import *
from tkinter import filedialog
import tkinter as tk
import xlwt
import openpyxl
import threading


def progress (rot):
    global PrgD
#    global pos
#    global rot
    rot[0] = ( rot[0] + 1 ) % 4
    PrgD.title ( "|/-\\"[ rot[0] ] )
#    rot[1].set ( "|/-\\"[ rot[0] ] )

def nullStr ( r, c ):
    global ws
    if ws.cell ( row=r, column=c ).value is None:
        return "null"
    else:
        return "\"" + str ( ws.cell(row=r, column=c).value ).replace ( "\r", "\\\\r" ).replace ( "\n", "\\\\n" ).replace ( "\"", "\\\"" ) + "\""

def nullNum ( r, c ):
    global ws
    if ws.cell ( row=r, column=c ).value is None:
        return 0
    else:
        return 1

def groupToJSoN ( grpID ):
    global ws
    week = "["
    c0 = startCols[grpID]
    c1 = c0 + 1
    c2 = c1 + 1
    c3 = c2 + 1
    for rN in range(4,74,12):
        week = week + "["
    #    day = [] Возм
        for h in range(1,12,2):
            r1 = rN + h
            r2 = r1 + 1
            f1 = nullNum ( r1, c0 )
            f2 = nullNum ( r2, c0 )
            if ws.cell( row=r1, column=c0 ).value == ws.cell( row=r2, column=c0).value:
                week = week + "{" +   \
                    "\"name\" : "    + nullStr( r1, c0 ) + "," + \
                    "\"type\" : "    + nullStr( r1, c1 ) + "," + \
                    "\"teacher\" : " + nullStr( r1, c2 ) + "," + \
                    "\"room\" : "    + nullStr( r1, c3 ) + "," + \
                    "\"week\" : null"  + \
                "},"
            else:
                week = week + "["
                if 0 != f1:
                    week = week + "{" + \
                            "\"name\" : "    + nullStr( r1, c0 ) + "," + \
                            "\"type\" : "    + nullStr( r1, c1 ) + "," + \
                            "\"teacher\" : " + nullStr( r1, c2 ) + "," + \
                            "\"room\" : "    + nullStr( r1, c3 ) + "," + \
                            "\"week\" : 1"   + \
                        "}"
                    if 0 != f2:
                        week = week + ","
                if 0 != f2:
                    week = week + "{"  + \
                        "\"name\" : "    + nullStr( r2, c0 ) + "," + \
                        "\"type\" : "    + nullStr( r2, c1 ) + "," + \
                        "\"teacher\" : " + nullStr( r2, c2 ) + "," + \
                        "\"room\" : "    + nullStr( r2, c3 ) + "," + \
                        "\"week\" : 2" + \
                    "}"
                week = week + "],"
        week = week.rstrip ( "," ) + "],"
    return week.rstrip(",") + "]"

def saveResult ():
    global LIST
    global groupNames
    idx = LIST.curselection()[0]
    grp = groupNames[idx]
    json = groupToJSoN ( idx )
    path = filedialog.asksaveasfilename ( title = "JSON File to Save", filetypes = [ ( "JSON Files", "*.json" ), ( "All Files", "*.*" ) ] )
    if not path:
        return
    if path[-5:].lower () != ".json":
        path = path + ".json"
    with open ( path, 'w' ) as file:
        file.write ( json )
        file.close ()


winW = 150
winH = 240


root1 = Tk()
lbl = Label(root1, text = "Пожалуйста подождите, \n загрузка...")
lbl.pack()
MAIN = Tk()
MAIN.resizable ( False, False )
winX = ( MAIN.winfo_screenwidth() - winW ) // 2
winY = ( MAIN.winfo_screenheight() - winH ) // 2
MAIN.geometry("{}x{}+{}+{}".format ( winW, winH, winX, winY ) )
#MAIN.withdraw ()

path = filedialog.askopenfilename (  title = "Choose File to Open", filetypes = [ ( "Microsoft Excel Files", [ "*.xls", "*.xlsx" ] ), ( "All Files", "*.*" ) ]  )

if not path:
    MAIN.destroy ()
    exit ()

if 0 == len ( path ):
    MAIN.destroy ()
    exit ()

#pos = 0
rot = []
rot.append ( 0 )
rot.append ( StringVar () )
rot[1].set ("|")
_pd = []
_pd.append ( rot )


MAIN.title('Выгрузка расписаний')
LIST = Listbox ( MAIN, selectmode=SINGLE)
LIST.pack ( pady=15 )
SAVE = Button ( MAIN, text="Выгрузить расписание", command=saveResult )
SAVE.pack ( pady=10 )

# PRogGess Dialog
#PrgD = Toplevel ( MAIN )
#PRGD = Tk ()
#PrgD.geometry ( "250x20+100+100" )
#PrgD.title ( "Пожалуйста, подождите" )
#PrgD.title ( "|" )
#Label ( PrgD, text = "Пожалуйста, подождите" ).pack ()
#skip =
#Label ( PrgD, textvariable = rot[1] ).pack ()
#PrgD.mainloop ()
#PrgD.deiconify ()

#tm = threading.Timer ( 1.0, progress, _pd )
#tm.start ()

wb = openpyxl.load_workbook(path)

#tm.cancel ()
#PrgD.destroy ()

# Main Window

ws = wb.active
#i = ws.max_row
#print (i)
w = ws.max_column
#print (w)

groupNames = []
startCols = []

for Ci in range(5,w,1):
  #print("" + str(Ci) + str( sheet.cell(row=2, column=Ci).value ))
  if ws.cell(row=3, column=Ci).value == "Предмет":
    gName = str( ws.cell(row=2, column=Ci).value )
    groupNames.append ( gName )
    startCols.append ( Ci )
    LIST.insert ( END, gName )

MAIN.deiconify ()
MAIN.mainloop ()
